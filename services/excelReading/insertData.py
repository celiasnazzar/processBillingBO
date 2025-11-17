from io import BytesIO
from fastapi import HTTPException
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from copy import copy


def insertData(data: dict, content: bytes) -> bytes:
    """Añade una nueva fila al Excel dentro de la tabla, manteniendo el formato."""

    # 1) Validar columnas con pandas (igual que antes)
    try:
        df = pd.read_excel(BytesIO(content), sheet_name="Tabla1", header=2)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo leer la hoja: {exc}")

    missing_cols = set(data.keys()) - set(df.columns)
    if missing_cols:
        raise HTTPException(
            status_code=400,
            detail=f"Columnas no encontradas: {missing_cols}. Disponibles: {list(df.columns)}"
        )

    # 2) Cargar workbook y hoja con openpyxl
    wb = load_workbook(BytesIO(content))
    ws = wb["Tabla1"]

    # 3) Localizar la tabla de Excel (ListObject)
    try:
        tbl = ws.tables["Tabla1"]
    except Exception:
        tbl = next((t for t in ws._tables if t.name == "Tabla1"), None)

    if tbl is None:
        raise HTTPException(status_code=400, detail="No se encontró la tabla de Excel 'Tabla1' en la hoja.")

    # 4) Rango actual de la tabla
    min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
    print(f"Rango tabla antes de inserción: {tbl.ref}, min_col={min_col}, min_row={min_row}, max_col={max_col}, max_row={max_row}")
    header_row = min_row    
    next_row = max_row + 1   

    # 5) Leer nombres de columna dentro del rango de la tabla
    column_names = []
    for col_idx in range(min_col, max_col + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        if cell.value is not None:
            column_names.append(str(cell.value).strip())
        else:
            column_names.append(None)

    # 6) Escribir los datos en la nueva fila dentro del rango de la tabla
    for offset, col_name in enumerate(column_names):
        col_idx = min_col + offset
        if col_name and col_name in data:
            cell = ws.cell(row=next_row, column=col_idx)
            cell.value = data[col_name]

            # Copiar estilo de la celda superior (última fila de la tabla anterior)
            source_cell = ws.cell(row=next_row - 1, column=col_idx)
            if source_cell.has_style:
                cell.font = copy(source_cell.font)
                cell.border = copy(source_cell.border)
                cell.fill = copy(source_cell.fill)
                cell.number_format = source_cell.number_format
                cell.alignment = copy(source_cell.alignment)

    # 7) Ampliar el rango de la tabla para incluir la nueva fila
    from openpyxl.utils import get_column_letter
    start_col_letter = get_column_letter(min_col)
    end_col_letter = get_column_letter(max_col)
    tbl.ref = f"{start_col_letter}{min_row}:{end_col_letter}{next_row}"

    # 8) Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
